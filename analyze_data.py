import pandas as pd
import openpyxl

df = pd.read_csv('output/us_correctional_facilities_master.csv')

print("=== DATA INTEGRITY REVIEW ===")
# duplicate facility names at same city+state
dups = df[df.duplicated(subset=['facility_name', 'city', 'state'], keep=False)]
print(f"Duplicate names at same city+state: {len(dups)}")

# empty street addresses
empty_addr = df[df['street_address'].isna() | (df['street_address'] == '')]
print(f"Empty street addresses: {len(empty_addr)}")

# anomalous facility names (pure numbers, single character names)
anom_names = df[df['facility_name'].str.match(r'^\d+$|^.$', na=False)]
print(f"Anomalous facility names: {len(anom_names)}")
if len(anom_names) > 0:
    print(anom_names[['facility_name']])

# suspiciously duplicate coordinates
dup_coords = df[df.duplicated(subset=['latitude', 'longitude'], keep=False)]
print(f"Duplicate coordinates count: {len(dup_coords)}")

# facilities in impossible US geography
# bounds check done by tests, but let's see if there are any that aren't US
out_lon = df[~(((df['longitude'] >= -180.0) & (df['longitude'] <= -64.0)) | ((df['longitude'] >= 144.0) & (df['longitude'] <= 146.0)))]
print(f"Impossible geography count: {len(out_lon)}")

# 20 random records from regions: just printing counts to make sure we can
regions = {
    'Northeast': ['CT', 'ME', 'MA', 'NH', 'RI', 'VT', 'NJ', 'NY', 'PA'],
    'Midwest': ['IL', 'IN', 'MI', 'OH', 'WI', 'IA', 'KS', 'MN', 'MO', 'NE', 'ND', 'SD'],
    'South': ['DE', 'FL', 'GA', 'MD', 'NC', 'SC', 'VA', 'DC', 'WV', 'AL', 'KY', 'MS', 'TN', 'AR', 'LA', 'OK', 'TX'],
    'West': ['AZ', 'CO', 'ID', 'MT', 'NV', 'NM', 'UT', 'WY', 'AK', 'CA', 'HI', 'OR', 'WA'],
    'Territories': ['PR', 'GU', 'VI', 'MP', 'AS']
}
for name, states in regions.items():
    print(f"{name} records: {len(df[df['state'].isin(states)])}")

# BOTH street_address AND phone_number AND website
full_contact = df[df['street_address'].notna() & (df['street_address'] != '') & 
                  df['phone_number'].notna() & (df['phone_number'] != '') & 
                  df['website'].notna() & (df['website'] != '')]
print(f"Full contact populated: {len(full_contact)}")

# pop > 3x capacity
df['design_capacity'] = pd.to_numeric(df['design_capacity'], errors='coerce')
df['population'] = pd.to_numeric(df['population'], errors='coerce')
high_pop = df[(df['population'] > 3 * df['design_capacity']) & (df['design_capacity'] > 0)]
print(f"Pop > 3x capacity count: {len(high_pop)}")
if len(high_pop) > 0:
    print(high_pop[['facility_name', 'design_capacity', 'population', 'data_source']].head(10))

# BOP-only no coords
bop_only = df[df['data_source'] == 'Federal Bureau of Prisons (BOP)']
bop_no_coords = bop_only[bop_only['latitude'].isna() | bop_only['longitude'].isna()]
print(f"BOP-only with no coords: {len(bop_no_coords)}")

print("\n=== EXCEL WORKBOOK REVIEW ===")
wb = openpyxl.load_workbook('output/us_correctional_facilities_master.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)

# Check ZIPs and FIPS for lost data type fidelity (stored as numbers instead of strings)
ws1 = wb['Master Facilities Directory']
zip_col_idx = 10
fips_col_idx = 12

numeric_zips = 0
numeric_fips = 0

for row in range(2, ws1.max_row + 1):
    zip_val = ws1.cell(row=row, column=zip_col_idx).value
    fips_val = ws1.cell(row=row, column=fips_col_idx).value
    
    if zip_val and isinstance(zip_val, (int, float)):
        numeric_zips += 1
    if fips_val and isinstance(fips_val, (int, float)):
        numeric_fips += 1

print(f"Numeric ZIPs: {numeric_zips}")
print(f"Numeric FIPS: {numeric_fips}")

