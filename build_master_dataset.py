#!/usr/bin/env python3
"""
Prison Data Aggregator - Master Pipeline (Audited & Robust)
Builds a unified, standardized database and spreadsheets (CSV + XLSX)
of all US correctional facilities across Federal, State, County, and Private sectors.
"""

import os
import sys
import json
import time
import re
import math
import urllib.request
import urllib.parse
from datetime import datetime, timezone
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_FILE = os.path.join(OUTPUT_DIR, "us_correctional_facilities_master.csv")
XLSX_FILE = os.path.join(OUTPUT_DIR, "us_correctional_facilities_master.xlsx")
SUMMARY_FILE = os.path.join(OUTPUT_DIR, "dataset_summary.json")

print("="*70)
print("US CORRECTIONAL FACILITIES MASTER AGGREGATION PIPELINE")
print("="*70)

# ----------------------------------------------------------------------
# 1. INGESTION HELPERS
# ----------------------------------------------------------------------
def fetch_arcgis_features(service_url, name):
    cache_path = os.path.join(DATA_DIR, f"{name}_raw.json")
    if os.path.exists(cache_path) and os.path.getsize(cache_path) > 100000:
        print(f"[+] Using cached data for {name} ({os.path.getsize(cache_path):,} bytes)")
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)

    print(f"[*] Querying {name} ArcGIS FeatureServer...")
    count_url = f"{service_url}/query?where=1=1&returnCountOnly=true&f=json"
    req = urllib.request.Request(count_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=15) as r:
        total_count = json.loads(r.read().decode("utf-8")).get("count", 0)
    print(f"  -> Total records available on server: {total_count:,}")

    offset = 0
    batch_size = 2000
    features = []
    
    while offset < total_count or total_count == 0:
        q_url = (
            f"{service_url}/query?where=1=1&outFields=*&returnGeometry=true"
            f"&outSR=4326&f=json&resultOffset={offset}&resultRecordCount={batch_size}"
        )
        req = urllib.request.Request(q_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            res = json.loads(r.read().decode("utf-8"))
            batch = res.get("features", [])
            if not batch:
                break
            features.extend(batch)
            print(f"     Downloaded {len(features):,} / {total_count:,} records...")
            if len(batch) < batch_size or not res.get("exceededTransferLimit", True):
                break
            offset += len(batch)

    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(features, f)
    print(f"[+] Successfully saved {len(features):,} raw features for {name}")
    return features

# 1. Fetch HIFLD Data
print("\n[1/6] Ingesting Homeland Infrastructure Foundation-Level Data (HIFLD)...")
hifld_url_primary = "https://services4.arcgis.com/DZmRnAEdOfXI200k/arcgis/rest/services/Prison_Points/FeatureServer/0"
hifld_raw = fetch_arcgis_features(hifld_url_primary, "hifld_primary")

# 2. Fetch Federal Bureau of Prisons (BOP) directory
print("\n[2/6] Ingesting Federal Bureau of Prisons (BOP) official directory...")
bop_raw = []
bop_cache = os.path.join(DATA_DIR, "bop_raw.json")
if os.path.exists(bop_cache) and os.path.getsize(bop_cache) > 1000:
    print(f"[+] Using cached BOP directory ({os.path.getsize(bop_cache):,} bytes)")
    with open(bop_cache, "r", encoding="utf-8") as f:
        bop_data = json.load(f)
        bop_raw = bop_data.get("Locations", []) if isinstance(bop_data, dict) else bop_data
else:
    try:
        bop_url = "https://www.bop.gov/PublicInfo/execute/locations?todo=query&output=json"
        req = urllib.request.Request(bop_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            bop_data = json.loads(r.read().decode("utf-8"))
            with open(bop_cache, "w", encoding="utf-8") as f:
                json.dump(bop_data, f)
            bop_raw = bop_data.get("Locations", [])
            print(f"[+] Ingested {len(bop_raw)} BOP locations from official API")
    except Exception as e:
        print(f"[-] Live BOP fetch failed: {e}")

# ----------------------------------------------------------------------
# 3. NORMALIZATION, PARSING & ENRICHMENT
# ----------------------------------------------------------------------
print(f"\n[3/6] Normalizing, Classifying, and Parsing records...")

SENTINEL_STRINGS = {
    "NOT AVAILABLE", "UNAVAILABLE", "NONE", "NULL", "-999", "-999.0", "N/A", "NA", "UNKNOWN", 
    "NOT APPLICABLE", "NOT REPORTED", "UNSPECIFIED", "", "-1", "-1--1", "-1-", "000-000-0000"
}

SENTINEL_INTS = {-999, -1, 99999}

ACRONYMS = {
    'USP', 'FCI', 'ADX', 'ADMAX', 'FDC', 'MDC', 'FMC', 'FPC', 'MCC', 'FCC', 'BOP', 'DOC', 
    'USMS', 'ICE', 'DHS', 'DJJ', 'DOJ', 'SCI', 'SD', 'CCF', 'CF', 'ASPC', 'RRM', 
    'RO', 'CI', 'MSTC', 'YDC', 'JDC', 'AJDC', 'CJCF', 'MCI', 'WSP', 'NER', 'SER', 'NCR', 
    'SCR', 'WXR', 'MXR', 'HQ', 'II', 'III', 'IV', 'VI', 'NW', 'NE', 'SW', 'SE', 'US', 'USA', 'CCFW'
}

# Explicit FIPS and County Lookup for standalone BOP and territory installations
STANDALONE_COUNTY_FIPS_MAP = {
    "BOP-LOF": ("Santa Barbara", "06083"),
    "BOP-LOM": ("Santa Barbara", "06083"),
    "BOP-CSC": ("Sacramento", "06067"),
    "BOP-CLB": ("Los Angeles", "06037"),
    "BOP-WXR": ("San Joaquin", "06077"),
    "BOP-DET": ("Arapahoe", "08005"),
    "BOP-BOP": ("District of Columbia", "11001"),
    "BOP-COX": ("Sumter", "12119"),
    "BOP-COR": ("Sumter", "12119"),
    "BOP-SER": ("Fulton", "13121"),
    "BOP-GLN": ("Glynn", "13127"),
    "BOP-CCH": ("DuPage", "17043"),
    "BOP-NCR": ("Wyandotte", "20209"),
    "BOP-CKC": ("Wyandotte", "20209"),
    "BOP-MXR": ("Anne Arundel", "24003"),
    "BOP-CBR": ("Baltimore City", "24510"),
    "BOP-CDT": ("Washtenaw", "26161"),
    "BOP-CMS": ("Hennepin", "27053"),
    "BOP-CST": ("St. Louis City", "29510"),
    "10006284": ("Saipan Municipality", "69110"),
    "10006540": ("Guam", "66010"),
    "10006165": ("Guam", "66010"),
    "10006164": ("Guam", "66010"),
    "10006539": ("Guam", "66010"),
    "10006163": ("St. Croix Island", "78010"),
    "10006166": ("St. Croix Island", "78010"),
    "BOP-YAM": ("Yazoo", "28163"),
    "BOP-CRL": ("Granville", "37077"),
    "BOP-CCN": ("Hamilton", "39061"),
    "BOP-ALX": ("Union", "42119"),
    "BOP-NER": ("Philadelphia", "42101"),
    "BOP-CPG": ("Allegheny", "42003"),
    "BOP-CNV": ("Davidson", "47037"),
    "BOP-GRA": ("Dallas", "48113"),
    "BOP-SCR": ("Dallas", "48113"),
    "BOP-CDA": ("Dallas", "48113"),
    "BOP-CSA": ("Bexar", "48029"),
    "BOP-CSE": ("King", "53033"),
    "BOP-CAT": ("Fulton", "13121"),
    "BOP-BMX": ("Jefferson", "48245"),
    "BOP-BUX": ("Granville", "37077"),
    "BOP-FLX": ("Fremont", "08043"),
    "BOP-FOX": ("St. Francis", "05123"),
    "BOP-HAX": ("Preston", "54077"),
    "BOP-LOX": ("Santa Barbara", "06083"),
    "BOP-CMM": ("Miami-Dade", "12086"),
    "BOP-CMY": ("Montgomery", "01101"),
    "BOP-CNK": ("New York", "36061"),
    "BOP-OAX": ("Allen", "22003"),
    "BOP-PEX": ("Prince George", "51149"),
    "BOP-CPA": ("Philadelphia", "42101"),
    "BOP-CPH": ("Maricopa", "04013"),
    "BOP-POX": ("Grant", "22043"),
    "BOP-THX": ("Vigo", "18167"),
    "BOP-TCX": ("Pima", "04019"),
    "BOP-VIX": ("San Bernardino", "06071"),
    "BOP-YAX": ("Yazoo", "28163"),
    "BOP-MRG": ("Monongalia", "54061"),
    "BOP-HER": ("Lassen", "06035")
}

def clean_text(val):
    if val is None or pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.upper() in SENTINEL_STRINGS or val_str in SENTINEL_STRINGS:
        return ""
    return val_str

def clean_zip(val):
    text = clean_text(val)
    if not text:
        return ""
    digits = re.sub(r'[^0-9]', '', text)
    if len(digits) == 5:
        return digits
    elif len(digits) == 9:
        return f"{digits[:5]}-{digits[5:]}"
    elif len(digits) == 4:
        return f"0{digits}"
    elif len(digits) == 3:
        return f"00{digits}"
    return text

def clean_fips(val, state=""):
    text = clean_text(val)
    if not text:
        return ""
    digits = re.sub(r'[^0-9]', '', text)
    # Fix known Pickens County Alabama FIPS typo in upstream HIFLD (10107 -> 01107)
    if state == "AL" and digits == "10107":
        digits = "01107"
    elif len(digits) == 4:
        digits = f"0{digits}"
    return digits

def clean_phone(val):
    text = clean_text(val)
    if not text or text in ["-1--1", "-1", "0"]:
        return ""
    digits = re.sub(r'\D', '', text)
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    elif len(digits) == 11 and digits.startswith('1'):
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    elif len(digits) > 10:
        ext = digits[10:]
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:10]} Ext {ext}"
    if len(digits) < 7:
        return ""
    return text

def clean_int(val):
    if val is None or pd.isna(val):
        return None
    try:
        f = float(val)
        if f in SENTINEL_INTS or f < 0:
            return None
        return int(round(f))
    except (ValueError, TypeError):
        return None

def clean_coord(val, is_lat=True):
    if val is None or pd.isna(val):
        return None
    try:
        f = float(val)
        if is_lat:
            if 13.0 <= f <= 72.0:  # Valid US latitude range
                return round(f, 6)
        else:
            # Valid US longitude range (Lower 48, AK, Territories, Aleutians up to +180)
            if (-180.0 <= f <= -64.0) or (144.0 <= f <= 180.0):
                return round(f, 6)
        return None
    except (ValueError, TypeError):
        return None

def normalize_gender(val):
    """Normalize gender values to documented schema: Male, Female, Co-ed, Not Specified."""
    text = clean_text(val).title()
    if not text:
        return "Not Specified"
    if text.lower() == "mixed":
        return "Co-ed"
    if text in ("Male", "Female", "Co-Ed", "Co-ed"):
        return text if text != "Co-Ed" else "Co-ed"
    return "Not Specified"

def format_title(text):
    if not text:
        return ""
    text = re.sub(r'\s*-\s*', ' - ', text)
    text = re.sub(r'\s*/\s*', ' / ', text)
    words = text.split()
    formatted = []
    for w in words:
        prefix, core, suffix = '', w, ''
        m = re.match(r'^([\(\"\'\-\/]*)(.*?)([\)\"\'\-\/\,\.]*)$', w)
        if m:
            prefix, core, suffix = m.groups()
        
        core_upper = core.upper()
        core_clean = re.sub(r'[^A-Za-z0-9]', '', core_upper)
        
        # Possessive check: e.g. Men's, Women's, Sheriff's, McDonald's
        if core.lower().endswith("'s"):
            base = core[:-2].capitalize()
            base = re.sub(r'^Mc([a-z])', lambda x: 'Mc' + x.group(1).upper(), base)
            base = re.sub(r"^O'([a-z])", lambda x: "O'" + x.group(1).upper(), base)
            core_formatted = f"{base}'s"
        elif core_clean in ACRONYMS:
            core_formatted = core_upper
        elif core.lower() in ['and', 'of', 'for', 'in', 'at', 'on', 'to', 'the', 'de', 'la', 'del']:
            core_formatted = core.lower()
        else:
            core_formatted = core.capitalize()
            # Smart capitalization for Mc and O' prefixes
            core_formatted = re.sub(r'^Mc([a-z])', lambda x: 'Mc' + x.group(1).upper(), core_formatted)
            core_formatted = re.sub(r"^O'([a-z])", lambda x: "O'" + x.group(1).upper(), core_formatted)
            
        formatted.append(f"{prefix}{core_formatted}{suffix}")
        
    res = ' '.join(formatted)
    if res and res[0].islower():
        res = res[0].upper() + res[1:]
    res = re.sub(r'\s+-\s+', ' - ', res)
    res = re.sub(r'\s+/\s+', ' / ', res)
    res = re.sub(r"\bO'([a-z])", lambda x: "O'" + x.group(1).upper(), res)
    return res

# A. Consolidate HIFLD by unique FACILITYID
hifld_dict = {}
multipart_consolidated_count = 0

for feat in hifld_raw:
    attrs = feat.get("attributes", {})
    fac_id = clean_text(attrs.get("FACILITYID"))
    if not fac_id:
        print(f"[WARN] Skipping record missing FACILITYID: {attrs.get('NAME')}")
        continue
    
    geom = feat.get("geometry", {})
    lat = clean_coord(attrs.get("LATITUDE") or geom.get("y"), is_lat=True)
    lon = clean_coord(attrs.get("LONGITUDE") or geom.get("x"), is_lat=False)
    
    if fac_id not in hifld_dict:
        hifld_dict[fac_id] = (attrs, lat, lon)
    else:
        multipart_consolidated_count += 1
        existing_attrs, ex_lat, ex_lon = hifld_dict[fac_id]
        if ex_lat is None and lat is not None:
            hifld_dict[fac_id] = (attrs, lat, lon)
        elif ex_lat is not None and lat is not None:
            # Retaining primary centroid while consolidating secondary boundary geometry
            pass

print(f"[+] Unique baseline physical facilities from HIFLD: {len(hifld_dict):,}")
print(f"[+] Multi-part polygon features consolidated: {multipart_consolidated_count:,} secondary boundary nodes merged (primary centroids preserved)")

# B. Enrich with BOP official data with TYPE-GUARDED & CAMP-PARITY NON-COLLISION MATCHING
bop_matched_hifld_ids = set()
enriched_hifld_count = 0
standalone_bop_records = []

for bop in bop_raw:
    bop_code = clean_text(bop.get("code")).upper()
    bop_name = clean_text(bop.get("nameTitle") or bop.get("nameDisplay") or bop.get("name")).upper()
    bop_st = clean_text(bop.get("state")).upper()
    bop_city = clean_text(bop.get("city")).upper()
    bop_type = clean_text(bop.get("type")).upper()
    
    # 1. Administrative headquarters, regional offices, and RRM field offices are distinct entities (ingest as standalone)
    if bop_type in ("RRM", "RO", "CO", "FCC", "TRN", "STAFF TRAINING ACADEMY"):
        standalone_bop_records.append(bop)
        continue
        
    matched_id = None
    
    # Priority 1: Exact Title or Exact BOP Code Match
    for fac_id, (attrs, lat, lon) in hifld_dict.items():
        if fac_id in bop_matched_hifld_ids:
            continue
        if clean_text(attrs.get("STATE")).upper() != bop_st:
            continue
            
        h_name = clean_text(attrs.get("NAME")).upper()
        h_type = clean_text(attrs.get("TYPE")).upper()
        
        is_fed = (h_type in ("FEDERAL", "MULTI") or "FEDERAL" in h_name or "USP " in h_name or 
                  "FCI " in h_name or "MDC " in h_name or "FDC " in h_name or "MCC " in h_name or 
                  "FMC " in h_name or "FPC " in h_name or "ADX " in h_name)
        if not is_fed:
            continue
            
        if bop_name == h_name or (bop_code and (f" {bop_code} " in f" {h_name} " or h_name.startswith(f"{bop_code} "))):
            matched_id = fac_id
            break

    # Priority 2: Core Institution Name & Camp-Parity Match
    if not matched_id:
        b_name_clean = re.sub(r'[^A-Z0-9]', '', bop_name)
        b_city_clean = re.sub(r'[^A-Z0-9]', '', bop_city)
        b_is_camp = ("CAMP" in bop_name) or ("FPC" in bop_name)
        
        for fac_id, (attrs, lat, lon) in hifld_dict.items():
            if fac_id in bop_matched_hifld_ids:
                continue
            if clean_text(attrs.get("STATE")).upper() != bop_st:
                continue
                
            h_name = clean_text(attrs.get("NAME")).upper()
            h_type = clean_text(attrs.get("TYPE")).upper()
            
            is_fed = (h_type in ("FEDERAL", "MULTI") or "FEDERAL" in h_name or "USP " in h_name or 
                      "FCI " in h_name or "MDC " in h_name or "FDC " in h_name or "MCC " in h_name or 
                      "FMC " in h_name or "FPC " in h_name or "ADX " in h_name)
            if not is_fed:
                continue
                
            h_name_clean = re.sub(r'[^A-Z0-9]', '', h_name)
            h_city_clean = re.sub(r'[^A-Z0-9]', '', clean_text(attrs.get("CITY")).upper())
            h_is_camp = ("CAMP" in h_name) or ("FPC" in h_name)
            
            # Enforce camp-to-camp matching parity (prevents camp from stealing parent USP/FCI URL)
            if b_is_camp != h_is_camp:
                continue
                
            b_core = re.sub(r'^(FCI|USP|FDC|MDC|MCC|FMC|FPC|ADX|FEDERALPRISONCAMP|FEDERALCORRECTIONALINSTITUTION|UNITEDSTATESPENITENTIARY)', '', b_name_clean)
            h_core = re.sub(r'^(FCI|USP|FDC|MDC|MCC|FMC|FPC|ADX|FEDERALPRISONCAMP|FEDERALCORRECTIONALINSTITUTION|UNITEDSTATESPENITENTIARY)', '', h_name_clean)
            
            if (b_city_clean == h_city_clean or not b_city_clean) and b_core and (b_core == h_core or b_core in h_core):
                matched_id = fac_id
                break

    if matched_id:
        bop_matched_hifld_ids.add(matched_id)
        enriched_hifld_count += 1
        attrs, lat, lon = hifld_dict[matched_id]
        
        if bop.get("url"):
            attrs["WEBSITE"] = f"https://www.bop.gov{bop.get('url')}"
        elif not attrs.get("WEBSITE") or "bop.gov" not in str(attrs.get("WEBSITE")).lower():
            attrs["WEBSITE"] = "https://www.bop.gov/locations/"
            
        if bop.get("phoneNumber"):
            attrs["TELEPHONE"] = bop.get("phoneNumber")
        if bop.get("securityLevel") and attrs.get("SECURELVL") in ["NOT AVAILABLE", "", None]:
            attrs["SECURELVL"] = bop.get("securityLevel").upper()
        if bop.get("gender") and attrs.get("GENDER") in ["NOT AVAILABLE", "", None]:
            attrs["GENDER"] = bop.get("gender").upper()
    else:
        standalone_bop_records.append(bop)

print(f"[+] HIFLD records enriched with official BOP contact/URLs: {enriched_hifld_count:,}")
print(f"[+] Standalone federal institutions / administrative offices added: {len(standalone_bop_records):,}")

# C. Build normalized master record list
master_records = []

# 1. Process all HIFLD unique records
for fac_id, (attrs, lat, lon) in hifld_dict.items():
    raw_name = clean_text(attrs.get("NAME"))
    state = clean_text(attrs.get("STATE")).upper()
    raw_type = clean_text(attrs.get("TYPE")).upper()
    raw_sec = clean_text(attrs.get("SECURELVL")).upper()
    raw_status = clean_text(attrs.get("STATUS")).upper()
    
    if raw_status in ["OPEN", "ACTIVE", "OPERATIONAL"]:
        status = "Open"
    elif raw_status in ["CLOSED", "INACTIVE", "DECOMMISSIONED"]:
        status = "Closed"
    else:
        status = "Not Available" if not raw_status else raw_status.title()

    if "FED" in raw_type or "FEDERAL" in raw_type or "BOP" in raw_name.upper():
        jurisdiction = "Federal"
    elif "STATE" in raw_type:
        jurisdiction = "State"
    elif "COUNTY" in raw_type or "JAIL" in raw_name.upper():
        jurisdiction = "County / Local"
    elif "LOCAL" in raw_type or "MUNICIPAL" in raw_type:
        jurisdiction = "Municipal / Local"
    elif "PRIVATE" in raw_type:
        jurisdiction = "Private"
    elif "MULTI" in raw_type:
        jurisdiction = "Multi-Jurisdiction"
    else:
        jurisdiction = "Not Specified"

    if "MAX" in raw_sec:
        security_level = "Maximum"
    elif "CLOSE" in raw_sec:
        security_level = "Close"
    elif "MED" in raw_sec:
        security_level = "Medium"
    elif "MIN" in raw_sec:
        security_level = "Minimum"
    elif "JUV" in raw_sec:
        security_level = "Juvenile"
    elif "MULTI" in raw_sec:
        security_level = "Multi-Level"
    elif "ADMIN" in raw_sec:
        security_level = "Administrative"
    else:
        security_level = "Not Specified"

    name_upper = raw_name.upper()
    if "JUVENILE" in name_upper or "YOUTH" in name_upper or security_level == "Juvenile":
        facility_type = "Juvenile Detention / Residential"
    elif "JAIL" in name_upper or "DETENTION CENTER" in name_upper or jurisdiction == "County / Local":
        facility_type = "County / Local Jail"
    elif "PRISON" in name_upper or "CORRECTIONAL INSTITUTION" in name_upper or "CORRECTIONAL FACILITY" in name_upper or "PENITENTIARY" in name_upper:
        facility_type = "State / Federal Prison"
    elif "COMMUNITY" in name_upper or "HALFWAY" in name_upper or "REENTRY" in name_upper or "WORK RELEASE" in name_upper:
        facility_type = "Community Corrections / Re-entry"
    elif "CAMP" in name_upper:
        facility_type = "Work Camp / Conservation Camp"
    elif "MEDICAL" in name_upper or "PSYCH" in name_upper or "HOSPITAL" in name_upper:
        facility_type = "Correctional Medical / Psychiatric"
    else:
        facility_type = "Adult Correctional Facility"

    cap = clean_int(attrs.get("CAPACITY"))
    pop = clean_int(attrs.get("POPULATION"))
    phone = clean_phone(attrs.get("TELEPHONE"))
    website = clean_text(attrs.get("WEBSITE") or attrs.get("SOURCE"))
    if not website.lower().startswith("http"):
        website = ""

    county_val = format_title(clean_text(attrs.get("COUNTY")))
    county_fips_val = clean_fips(attrs.get("COUNTYFIPS"), state=state)
    
    # Impute territory/missing county if in lookup
    if fac_id in STANDALONE_COUNTY_FIPS_MAP:
        county_val, county_fips_val = STANDALONE_COUNTY_FIPS_MAP[fac_id]

    rec = {
        "facility_id": fac_id,
        "facility_name": format_title(raw_name),
        "jurisdiction": jurisdiction,
        "facility_type": facility_type,
        "security_level": security_level,
        "operational_status": status,
        "street_address": format_title(clean_text(attrs.get("ADDRESS"))),
        "city": format_title(clean_text(attrs.get("CITY"))),
        "state": state,
        "zip_code": clean_zip(attrs.get("ZIP")),
        "county": county_val,
        "county_fips": county_fips_val,
        "phone_number": phone,
        "website": website,
        "latitude": lat,
        "longitude": lon,
        "design_capacity": cap,
        "population": pop,
        "gender": normalize_gender(attrs.get("GENDER")),
        "data_source": "DHS HIFLD Critical Infrastructure"
    }
    master_records.append(rec)

# 2. Process Standalone BOP records
for bop in standalone_bop_records:
    bop_name = clean_text(bop.get("nameTitle") or bop.get("nameDisplay") or bop.get("name"))
    if not bop_name:
        continue

    lat = clean_coord(bop.get("latitude"), is_lat=True)
    lon = clean_coord(bop.get("longitude"), is_lat=False)
    bop_type_desc = clean_text(bop.get("faclTypeDescription") or bop.get("type"))
    bop_sec = clean_text(bop.get("securityLevel")).title() or "Administrative"
    bop_gender = normalize_gender(bop.get("gender"))
    bop_code = clean_text(bop.get("code")).upper()
    bop_url = "https://www.bop.gov" + bop.get("url") if bop.get("url") else "https://www.bop.gov/locations/"
    bop_st = clean_text(bop.get("state")).upper()
    
    fac_id_key = f"BOP-{bop_code}" if bop_code else f"BOP-{len(master_records)}"
    county_val, county_fips_val = STANDALONE_COUNTY_FIPS_MAP.get(fac_id_key, ("", ""))

    rec = {
        "facility_id": fac_id_key,
        "facility_name": format_title(bop_name),
        "jurisdiction": "Federal",
        "facility_type": format_title(bop_type_desc) if bop_type_desc else "Federal Bureau of Prisons (BOP)",
        "security_level": bop_sec,
        "operational_status": "Open",
        "street_address": format_title(clean_text(bop.get("address"))),
        "city": format_title(clean_text(bop.get("city"))),
        "state": bop_st,
        "zip_code": clean_zip(bop.get("zipCode")),
        "county": county_val,
        "county_fips": county_fips_val,
        "phone_number": clean_phone(bop.get("phoneNumber")),
        "website": bop_url,
        "latitude": lat,
        "longitude": lon,
        "design_capacity": clean_int(bop.get("capacity")),
        "population": clean_int(bop.get("population")),
        "gender": bop_gender,
        "data_source": "Federal Bureau of Prisons (BOP)"
    }
    master_records.append(rec)

# ----------------------------------------------------------------------
# 4. FINAL SORTING & STRUCTURE
# ----------------------------------------------------------------------
print(f"\n[4/6] Structuring and indexing master dataset...")

df_master = pd.DataFrame(master_records)

# Sort cleanly by State, Jurisdiction, City, and Facility Name
df_master = df_master.sort_values(by=['state', 'jurisdiction', 'city', 'facility_name']).reset_index(drop=True)

# Convert numeric capacity and population to Nullable Int64 so CSV exports clean integers without .0
df_master['design_capacity'] = df_master['design_capacity'].astype('Int64')
df_master['population'] = df_master['population'].astype('Int64')

print(f"[+] Total master facilities consolidated: {len(df_master):,}")

# ----------------------------------------------------------------------
# 5. EXPORTING MASTER SPREADSHEETS (CSV & EXCEL)
# ----------------------------------------------------------------------
print(f"\n[5/6] Exporting CSV and Multi-Tab Excel Spreadsheets...")

# 1. Save CSV (Clean integer formatting, leading 0s intact)
df_master.to_csv(CSV_FILE, index=False, encoding='utf-8')
print(f"[+] Master CSV written: {CSV_FILE} ({os.path.getsize(CSV_FILE):,} bytes)")

# 2. Build Multi-Tab Master Excel (.xlsx)
wb = openpyxl.Workbook()

# Sheet 1: Master Directory
ws1 = wb.active
ws1.title = "Master Facilities Directory"
ws1.views.sheetView[0].showGridLines = True

font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

columns_config = [
    ("Facility ID", "facility_id", 14, align_center, "@"),
    ("Facility Name", "facility_name", 36, align_left, "@"),
    ("Jurisdiction", "jurisdiction", 18, align_center, "@"),
    ("Facility Classification", "facility_type", 28, align_left, "@"),
    ("Security Level", "security_level", 16, align_center, "@"),
    ("Operational Status", "operational_status", 14, align_center, "@"),
    ("Street Address", "street_address", 32, align_left, "@"),
    ("City", "city", 20, align_left, "@"),
    ("State", "state", 8, align_center, "@"),
    ("ZIP Code", "zip_code", 12, align_center, "@"),
    ("County", "county", 18, align_left, "@"),
    ("County FIPS", "county_fips", 12, align_center, "@"),
    ("Phone Number", "phone_number", 16, align_center, "@"),
    ("Website", "website", 30, align_left, "@"),
    ("Latitude", "latitude", 13, align_right, "0.000000"),
    ("Longitude", "longitude", 13, align_right, "0.000000"),
    ("Design Capacity", "design_capacity", 16, align_right, "#,##0"),
    ("Population", "population", 14, align_right, "#,##0"),
    ("Gender", "gender", 14, align_center, "@"),
    ("Data Source", "data_source", 28, align_left, "@")
]

# Write Sheet 1 Headers
ws1.row_dimensions[1].height = 28
for col_idx, (label, field, width, alignment, num_fmt) in enumerate(columns_config, start=1):
    cell = ws1.cell(row=1, column=col_idx, value=label)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# Write Sheet 1 Data
records_dict = df_master.to_dict(orient="records")
for row_idx, r in enumerate(records_dict, start=2):
    is_zebra = (row_idx % 2 == 0)
    for col_idx, (label, field, width, alignment, num_fmt) in enumerate(columns_config, start=1):
        val = r.get(field)
        cell = ws1.cell(row=row_idx, column=col_idx)
        if pd.isna(val) or val is None:
            cell.value = ""
        else:
            if field in ['design_capacity', 'population']:
                cell.value = int(val)
            else:
                cell.value = val
            cell.number_format = num_fmt
        cell.alignment = alignment
        cell.border = thin_border
        if is_zebra:
            cell.fill = fill_zebra
    if row_idx % 2500 == 0:
        print(f"     Formatted {row_idx:,} / {len(df_master):,} rows in Excel...")

ws1.freeze_panes = "A2"

# ----------------------------------------------------------------------
# Sheet 2: Summary by State
# ----------------------------------------------------------------------
ws2 = wb.create_sheet(title="State Summary")
ws2.views.sheetView[0].showGridLines = True

state_agg = df_master.groupby('state').agg(
    Total_Facilities=('facility_name', 'count'),
    Total_Capacity=('design_capacity', lambda x: int(x.dropna().sum())),
    Total_Population=('population', lambda x: int(x.dropna().sum())),
    Federal=('jurisdiction', lambda x: (x == 'Federal').sum()),
    State_DOC=('jurisdiction', lambda x: (x == 'State').sum()),
    County_Local=('jurisdiction', lambda x: (x == 'County / Local').sum()),
    Municipal_Local=('jurisdiction', lambda x: (x == 'Municipal / Local').sum()),
    Private=('jurisdiction', lambda x: (x == 'Private').sum()),
    Multi_Jurisdiction=('jurisdiction', lambda x: (x == 'Multi-Jurisdiction').sum()),
    Not_Specified=('jurisdiction', lambda x: (x == 'Not Specified').sum())
).reset_index().sort_values(by='Total_Facilities', ascending=False)

state_headers = [
    ("State / Terr", 12, align_center, "@"),
    ("Total Facilities", 16, align_right, "#,##0"),
    ("Reported Capacity", 18, align_right, "#,##0"),
    ("Reported Population", 18, align_right, "#,##0"),
    ("Federal Facilities", 18, align_right, "#,##0"),
    ("State DOC Facilities", 18, align_right, "#,##0"),
    ("County / Local Jails", 18, align_right, "#,##0"),
    ("Municipal Jails", 16, align_right, "#,##0"),
    ("Private Facilities", 16, align_right, "#,##0"),
    ("Multi-Jurisdiction", 16, align_right, "#,##0"),
    ("Not Specified", 14, align_right, "#,##0")
]

ws2.row_dimensions[1].height = 28
fill_state_header = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
for col_idx, (label, width, alignment, num_fmt) in enumerate(state_headers, start=1):
    c = ws2.cell(row=1, column=col_idx, value=label)
    c.font = font_header
    c.fill = fill_state_header
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, r in enumerate(state_agg.to_dict(orient="records"), start=2):
    vals = [
        (r['state'], align_center, "@"),
        (r['Total_Facilities'], align_right, "#,##0"),
        (r['Total_Capacity'], align_right, "#,##0"),
        (r['Total_Population'], align_right, "#,##0"),
        (r['Federal'], align_right, "#,##0"),
        (r['State_DOC'], align_right, "#,##0"),
        (r['County_Local'], align_right, "#,##0"),
        (r['Municipal_Local'], align_right, "#,##0"),
        (r['Private'], align_right, "#,##0"),
        (r['Multi_Jurisdiction'], align_right, "#,##0"),
        (r['Not_Specified'], align_right, "#,##0")
    ]
    for c_idx, (v, al, nf) in enumerate(vals, start=1):
        cell = ws2.cell(row=row_idx, column=c_idx, value=v)
        cell.alignment = al
        cell.number_format = nf
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = fill_zebra

ws2.freeze_panes = "A2"

# ----------------------------------------------------------------------
# Sheet 3: Summary by Jurisdiction
# ----------------------------------------------------------------------
ws3 = wb.create_sheet(title="Jurisdiction Summary")
ws3.views.sheetView[0].showGridLines = True

jur_agg = df_master.groupby(['jurisdiction', 'facility_type']).agg(
    Facility_Count=('facility_name', 'count'),
    Total_Capacity=('design_capacity', lambda x: int(x.dropna().sum())),
    Total_Population=('population', lambda x: int(x.dropna().sum())),
    Facilities_With_GPS=('latitude', lambda x: int(x.dropna().count()))
).reset_index().sort_values(by=['jurisdiction', 'Facility_Count'], ascending=[True, False])

jur_headers = [
    ("Jurisdiction", 22, align_left),
    ("Facility Classification", 32, align_left),
    ("Facility Count", 16, align_right),
    ("Reported Capacity", 18, align_right),
    ("Reported Population", 18, align_right),
    ("Mapped Coordinates", 18, align_right)
]

ws3.row_dimensions[1].height = 28
fill_jur_header = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
for col_idx, (label, width, alignment) in enumerate(jur_headers, start=1):
    c = ws3.cell(row=1, column=col_idx, value=label)
    c.font = font_header
    c.fill = fill_jur_header
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, r in enumerate(jur_agg.to_dict(orient="records"), start=2):
    ws3.cell(row=row_idx, column=1, value=r['jurisdiction']).alignment = align_left
    ws3.cell(row=row_idx, column=2, value=r['facility_type']).alignment = align_left
    
    for c_idx, val in enumerate([r['Facility_Count'], r['Total_Capacity'], r['Total_Population'], r['Facilities_With_GPS']], start=3):
        c = ws3.cell(row=row_idx, column=c_idx, value=val)
        c.alignment = align_right
        c.number_format = "#,##0"

    for c_idx in range(1, 7):
        ws3.cell(row=row_idx, column=c_idx).border = thin_border
        if row_idx % 2 == 0:
            ws3.cell(row=row_idx, column=c_idx).fill = fill_zebra

ws3.freeze_panes = "A2"

# ----------------------------------------------------------------------
# Sheet 4: Data Dictionary (Explicit 3-Column Mapping)
# ----------------------------------------------------------------------
ws4 = wb.create_sheet(title="Data Dictionary")
ws4.views.sheetView[0].showGridLines = True

dict_entries = [
    ("Facility ID", "facility_id", "Unique alphanumeric identifier assigned to each facility (HIFLD ID or BOP Code)."),
    ("Facility Name", "facility_name", "Official name of the correctional facility, standardized with Title Casing, Irish/Scottish prefix preservation (Mc/O'), and uppercase acronyms."),
    ("Jurisdiction", "jurisdiction", "Level of government authority: Federal, State, County / Local, Municipal / Local, Private, Multi-Jurisdiction, or Not Specified."),
    ("Facility Classification", "facility_type", "Operational facility classification (State / Federal Prison, County / Local Jail, Juvenile Detention, Community Corrections, Medical/Psychiatric, etc.)."),
    ("Security Level", "security_level", "Security classification: Maximum, Close, Medium, Minimum, Juvenile, Multi-Level, Administrative, Not Specified."),
    ("Operational Status", "operational_status", "Current operational status: Open, Closed, or Not Available."),
    ("Street Address", "street_address", "Standardized physical street address of the facility."),
    ("City", "city", "City where the facility is physically located."),
    ("State", "state", "Two-letter US postal state / territory abbreviation (all 50 states, DC, PR, GU, VI, MP)."),
    ("ZIP Code", "zip_code", "5-digit or 9-digit postal ZIP code with preserved leading zeroes."),
    ("County", "county", "County, parish, or borough name."),
    ("County FIPS", "county_fips", "5-digit Federal Information Processing Standard (FIPS) county code with preserved leading zeroes."),
    ("Phone Number", "phone_number", "Formatted telephone contact number for the facility (including extensions where reported)."),
    ("Website", "website", "Official website or government portal link for the facility."),
    ("Latitude", "latitude", "WGS84 decimal degrees latitude (North)."),
    ("Longitude", "longitude", "WGS84 decimal degrees longitude (West / East)."),
    ("Design Capacity", "design_capacity", "Official design or rated bed capacity of the facility."),
    ("Population", "population", "Reported inmate population count (valid zero population preserved for unoccupied/holding facilities)."),
    ("Gender", "gender", "Inmate gender housing designation: Male, Female, Co-ed, or Not Specified."),
    ("Data Source", "data_source", "Primary origin of data: DHS HIFLD Critical Infrastructure or Federal Bureau of Prisons.")
]

ws4.row_dimensions[1].height = 28
ws4.column_dimensions['A'].width = 26
ws4.column_dimensions['B'].width = 24
ws4.column_dimensions['C'].width = 85

fill_dict_header = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid")
c1 = ws4.cell(row=1, column=1, value="Display Column Header")
c2 = ws4.cell(row=1, column=2, value="CSV Field Name (snake_case)")
c3 = ws4.cell(row=1, column=3, value="Description & Definition")

for c in [c1, c2, c3]:
    c.font = font_header
    c.fill = fill_dict_header
    c.alignment = Alignment(horizontal="left", vertical="center")

for row_idx, (disp_name, f_name, desc) in enumerate(dict_entries, start=2):
    cell_disp = ws4.cell(row=row_idx, column=1, value=disp_name)
    cell_disp.font = Font(name="Calibri", size=11, bold=True)
    cell_disp.alignment = align_left
    cell_disp.border = thin_border
    
    cell_name = ws4.cell(row=row_idx, column=2, value=f_name)
    cell_name.font = Font(name="Consolas", size=10, color="203764")
    cell_name.alignment = align_left
    cell_name.border = thin_border
    
    cell_desc = ws4.cell(row=row_idx, column=3, value=desc)
    cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_desc.border = thin_border
    if row_idx % 2 == 0:
        cell_disp.fill = fill_zebra
        cell_name.fill = fill_zebra
        cell_desc.fill = fill_zebra

ws4.freeze_panes = "A2"

# Save workbook
wb.save(XLSX_FILE)
print(f"[+] Multi-Tab Master Excel Workbook saved: {XLSX_FILE} ({os.path.getsize(XLSX_FILE):,} bytes)")

# ----------------------------------------------------------------------
# 6. SUMMARY & VALIDATION AUDIT
# ----------------------------------------------------------------------
print(f"\n[6/6] Generating Validation Summary & Audit...")

summary = {
    "generated_at": datetime.now(timezone.utc).isoformat(),
    "total_unique_facilities": len(df_master),
    "total_reported_capacity": int(df_master['design_capacity'].dropna().sum()),
    "total_reported_population": int(df_master['population'].dropna().sum()),
    "facilities_with_gps_coordinates": int((df_master['latitude'].notna() & df_master['longitude'].notna()).sum()),
    "facilities_with_street_address": int((df_master['street_address'] != "").sum()),
    "facilities_with_phone_number": int((df_master['phone_number'] != "").sum()),
    "facilities_with_website": int((df_master['website'] != "").sum()),
    "states_and_territories_count": int(df_master['state'].nunique()),
    "states_list": sorted([s for s in df_master['state'].unique() if s]),
    "jurisdiction_breakdown": df_master['jurisdiction'].value_counts().to_dict(),
    "operational_status_breakdown": df_master['operational_status'].value_counts().to_dict(),
    "security_level_breakdown": df_master['security_level'].value_counts().to_dict(),
    "top_10_states_by_count": df_master['state'].value_counts().head(10).to_dict(),
    "output_files": {
        "master_csv": CSV_FILE,
        "master_excel": XLSX_FILE,
        "summary_json": SUMMARY_FILE
    }
}

with open(SUMMARY_FILE, "w", encoding="utf-8") as f:
    json.dump(summary, f, indent=2)

print("\n" + "="*70)
print("PIPELINE COMPLETED SUCCESSFULLY!")
print("="*70)
print(f"Total Master Facilities:        {summary['total_unique_facilities']:,}")
print(f"Mapped Coordinates (GPS):       {summary['facilities_with_gps_coordinates']:,} ({summary['facilities_with_gps_coordinates']/summary['total_unique_facilities']*100:.1f}%)")
print(f"Street Addresses:               {summary['facilities_with_street_address']:,} ({summary['facilities_with_street_address']/summary['total_unique_facilities']*100:.1f}%)")
print(f"Phone Numbers:                  {summary['facilities_with_phone_number']:,}")
print(f"Total Design Bed Capacity:      {summary['total_reported_capacity']:,}")
print(f"States & Territories Covered:   {summary['states_and_territories_count']} ({', '.join(summary['states_list'][:10])}...)")
print(f"\nFiles Generated:")
print(f"  • CSV:   {CSV_FILE}")
print(f"  • Excel: {XLSX_FILE}")
print(f"  • Audit: {SUMMARY_FILE}")
print("="*70)
